# ===================================================================
# Server Asset Management System (SAMS)
# 服务器资产管理系统
# Tech: Python / Streamlit / SQLite / pandas / openpyxl
# ===================================================================

import streamlit as st
import sqlite3
import hashlib
import pandas as pd
import os
import re
from datetime import datetime
from io import BytesIO

# ===================================================================
# Config
# ===================================================================

DB_PATH = os.path.join(os.environ.get("SAMS_DATA_DIR", os.path.dirname(os.path.abspath(__file__))), "sams.db")
PAGE_SIZE = 20

STATUS_COLORS = {
    "running": "#10b981",
    "stopped": "#f59e0b",
    "maintenance": "#3b82f6",
    "decommissioned": "#ef4444",
}

STATUS_LABELS = {
    "running": "运行中",
    "stopped": "已停止",
    "maintenance": "维护中",
    "decommissioned": "已下线",
}

STATUS_OPTIONS = ["running", "stopped", "maintenance", "decommissioned"]
OS_OPTIONS = ["Linux", "Windows", "macOS", "Other"]
RUNTIME_OPTIONS = ["physical", "vm", "docker", "k8s", "other"]

TEMPLATE_COLUMNS = [
    "hostname", "private_ip", "public_ip",
    "credentials", "credential_ref",
    "os_type", "os_version", "kernel_version",
    "cpu_cores", "memory_gb", "disk_info",
    "location", "business", "owner", "status",
    "purchase_date", "warranty_expire", "business_service",
    "app_framework", "db_info", "runtime_type", "runtime_detail",
    "port_info", "remarks",
]

COLUMN_LABELS_CN = {
    "hostname": "主机名",
    "private_ip": "内网IP",
    "public_ip": "公网IP",
    "credentials": "系统密码",
    "credential_ref": "应用密码",
    "os_type": "操作系统",
    "os_version": "系统版本",
    "kernel_version": "内核版本",
    "cpu_cores": "CPU(核)",
    "memory_gb": "内存(GB)",
    "disk_info": "磁盘信息",
    "location": "位置",
    "business": "业务线",
    "owner": "负责人",
    "status": "状态",
    "purchase_date": "采购日期",
    "warranty_expire": "保修截止",
    "business_service": "业务服务",
    "app_framework": "应用框架",
    "db_info": "数据库",
    "runtime_type": "运行方式",
    "runtime_detail": "运行详情",
    "port_info": "端口信息",
    "remarks": "备注",
}

COLUMN_LABELS_EN = {v: k for k, v in COLUMN_LABELS_CN.items()}

# ===================================================================
# Database Module
# ===================================================================

@st.cache_resource
def get_connection():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS servers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hostname TEXT NOT NULL,
            private_ip TEXT NOT NULL UNIQUE,
            public_ip TEXT DEFAULT '',
            os_type TEXT DEFAULT '',
            os_version TEXT DEFAULT '',
            kernel_version TEXT DEFAULT '',
            cpu_cores INTEGER DEFAULT 0,
            memory_gb INTEGER DEFAULT 0,
            disk_info TEXT DEFAULT '',
            location TEXT DEFAULT '',
            business TEXT DEFAULT '',
            owner TEXT DEFAULT '',
            status TEXT DEFAULT 'running',
            purchase_date TEXT DEFAULT '',
            warranty_expire TEXT DEFAULT '',
            business_service TEXT DEFAULT '',
            app_framework TEXT DEFAULT '',
            db_info TEXT DEFAULT '',
            runtime_type TEXT DEFAULT '',
            runtime_detail TEXT DEFAULT '',
            port_info TEXT DEFAULT '',
            credentials TEXT DEFAULT '',
            credential_ref TEXT DEFAULT '',
            remarks TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT DEFAULT (datetime('now', 'localtime'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        )
    """)
    conn.commit()
    create_indexes()
    _create_default_admin()


def create_indexes():
    conn = get_connection()
    indexes = [
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_private_ip ON servers(private_ip)",
        "CREATE INDEX IF NOT EXISTS idx_hostname ON servers(hostname)",
        "CREATE INDEX IF NOT EXISTS idx_owner ON servers(owner)",
        "CREATE INDEX IF NOT EXISTS idx_business ON servers(business)",
        "CREATE INDEX IF NOT EXISTS idx_status ON servers(status)",
        "CREATE INDEX IF NOT EXISTS idx_os_type ON servers(os_type)",
        "CREATE INDEX IF NOT EXISTS idx_updated_at ON servers(updated_at)",
    ]
    for idx_sql in indexes:
        conn.execute(idx_sql)
    conn.commit()


def _create_default_admin():
    conn = get_connection()
    existing = conn.execute("SELECT id FROM users WHERE username = ?", ("admin",)).fetchone()
    if not existing:
        pw_hash = hashlib.sha256("admin123".encode()).hexdigest()
        conn.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
            ("admin", pw_hash, "admin"),
        )
        conn.commit()


# ===================================================================
# Security Module
# ===================================================================

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()


def verify_password(password: str, password_hash: str) -> bool:
    return hash_password(password) == password_hash


# ===================================================================
# Auth Module
# ===================================================================

def check_auth() -> bool:
    return st.session_state.get("authenticated", False)


def login(username: str, password: str) -> bool:
    conn = get_connection()
    user = conn.execute(
        "SELECT * FROM users WHERE username = ?", (username,)
    ).fetchone()
    if user and verify_password(password, user["password_hash"]):
        st.session_state.authenticated = True
        st.session_state.username = user["username"]
        st.session_state.user_role = user["role"]
        return True
    return False


def logout():
    for key in list(st.session_state.keys()):
        del st.session_state[key]


# ===================================================================
# CRUD Module
# ===================================================================

def add_server(data: dict) -> int:
    conn = get_connection()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    columns = [
        "hostname", "private_ip", "public_ip",
        "credentials", "credential_ref",
        "os_type", "os_version", "kernel_version",
        "cpu_cores", "memory_gb", "disk_info",
        "location", "business", "owner", "status",
        "purchase_date", "warranty_expire", "business_service",
        "app_framework", "db_info", "runtime_type", "runtime_detail",
        "port_info", "remarks",
    ]
    placeholders = ", ".join(["?"] * len(columns))
    cols_str = ", ".join(columns)
    values = [data.get(c, "") for c in columns]
    cursor = conn.execute(
        f"INSERT INTO servers ({cols_str}, created_at, updated_at) VALUES ({placeholders}, ?, ?)",
        values + [now, now],
    )
    conn.commit()
    return cursor.lastrowid


def update_server(server_id: int, data: dict) -> bool:
    conn = get_connection()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    columns = [
        "hostname", "private_ip", "public_ip", "os_type", "os_version",
        "kernel_version", "cpu_cores", "memory_gb", "disk_info",
        "location", "business", "owner", "status",
        "purchase_date", "warranty_expire", "business_service",
        "app_framework", "db_info", "runtime_type", "runtime_detail",
        "port_info", "credentials", "credential_ref", "remarks",
    ]
    set_clause = ", ".join([f"{c} = ?" for c in columns])
    values = [data.get(c, "") for c in columns] + [now, server_id]
    conn.execute(
        f"UPDATE servers SET {set_clause}, updated_at = ? WHERE id = ?", values
    )
    conn.commit()
    return True


def delete_server(server_id: int):
    conn = get_connection()
    conn.execute("DELETE FROM servers WHERE id = ?", (server_id,))
    conn.commit()


def get_server_by_id(server_id: int) -> dict | None:
    conn = get_connection()
    row = conn.execute("SELECT * FROM servers WHERE id = ?", (server_id,)).fetchone()
    return dict(row) if row else None


def get_servers(
    page: int = 1,
    search: str = "",
    status_filter: list | None = None,
    os_filter: list | None = None,
    runtime_filter: list | None = None,
    business_filter: list | None = None,
    sort_by: str = "updated_at",
    sort_order: str = "DESC",
) -> tuple[list, int]:
    conn = get_connection()
    conditions = []
    params = []

    if search:
        search_cols = ["hostname", "private_ip", "owner", "business"]
        search_clauses = " OR ".join([f"{c} LIKE ?" for c in search_cols])
        conditions.append(f"({search_clauses})")
        params += [f"%{search}%"] * len(search_cols)

    if status_filter:
        conditions.append(f"status IN ({','.join(['?']*len(status_filter))})")
        params += status_filter

    if os_filter:
        conditions.append(f"os_type IN ({','.join(['?']*len(os_filter))})")
        params += os_filter

    if runtime_filter:
        conditions.append(f"runtime_type IN ({','.join(['?']*len(runtime_filter))})")
        params += runtime_filter

    if business_filter:
        conditions.append(f"business IN ({','.join(['?']*len(business_filter))})")
        params += business_filter

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    valid_sort_cols = ["hostname", "private_ip", "os_type", "cpu_cores", "memory_gb",
                       "business", "owner", "status", "updated_at", "created_at"]
    if sort_by not in valid_sort_cols:
        sort_by = "updated_at"
    if sort_order.upper() not in ("ASC", "DESC"):
        sort_order = "DESC"

    count_row = conn.execute(
        f"SELECT COUNT(*) FROM servers {where}", params
    ).fetchone()
    total = count_row[0]

    offset = (page - 1) * PAGE_SIZE
    rows = conn.execute(
        f"SELECT * FROM servers {where} ORDER BY {sort_by} {sort_order} LIMIT ? OFFSET ?",
        params + [PAGE_SIZE, offset],
    ).fetchall()

    return [dict(r) for r in rows], total


def get_distinct_values(column: str) -> list:
    conn = get_connection()
    rows = conn.execute(
        f"SELECT DISTINCT {column} FROM servers WHERE {column} IS NOT NULL AND {column} != '' ORDER BY {column}"
    ).fetchall()
    return [r[column] for r in rows]


# ===================================================================
# Statistics Module
# ===================================================================

@st.cache_data(ttl=60)
def get_dashboard_stats() -> dict:
    conn = get_connection()
    stats = {}

    stats["total"] = conn.execute("SELECT COUNT(*) FROM servers").fetchone()[0]

    for status in STATUS_OPTIONS:
        stats[status] = conn.execute(
            "SELECT COUNT(*) FROM servers WHERE status = ?", (status,)
        ).fetchone()[0]

    os_rows = conn.execute(
        "SELECT os_type, COUNT(*) as cnt FROM servers WHERE os_type != '' GROUP BY os_type ORDER BY cnt DESC"
    ).fetchall()
    stats["os_distribution"] = {r["os_type"]: r["cnt"] for r in os_rows}

    runtime_rows = conn.execute(
        "SELECT runtime_type, COUNT(*) as cnt FROM servers WHERE runtime_type != '' GROUP BY runtime_type ORDER BY cnt DESC"
    ).fetchall()
    stats["runtime_distribution"] = {r["runtime_type"]: r["cnt"] for r in runtime_rows}

    business_rows = conn.execute(
        "SELECT business, COUNT(*) as cnt FROM servers WHERE business != '' GROUP BY business ORDER BY cnt DESC"
    ).fetchall()
    stats["business_distribution"] = {r["business"]: r["cnt"] for r in business_rows}

    return stats


def get_db_size() -> str:
    if os.path.exists(DB_PATH):
        size_bytes = os.path.getsize(DB_PATH)
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.1f} KB"
        else:
            return f"{size_bytes / (1024 * 1024):.1f} MB"
    return "N/A"


# ===================================================================
# Excel Module
# ===================================================================

def generate_template(with_data: bool = False) -> BytesIO:
    from openpyxl.styles import Font, PatternFill
    cn_columns = [COLUMN_LABELS_CN[c] for c in TEMPLATE_COLUMNS]
    output = BytesIO()

    if with_data:
        sample_rows = [
            # hostname, private_ip, public_ip, credentials, credential_ref, os_type, os_version, kernel_version, cpu, mem, disk, location, business, owner, status, purchase_date, warranty_expire, business_service, app_framework, db_info, runtime_type, runtime_detail, port_info, remarks
            ["web-prod-01", "10.0.1.10", "203.0.113.10", "", "keycloak-vault-01", "Linux", "Ubuntu 22.04", "5.15.0-91-generic", 8, 32, "500GB SSD",
             "北京-亦庄-A区", "支付系统", "张三", "running", "2025-01-15", "2028-01-15",
             "支付网关服务", "Spring Boot", "MySQL 8.0", "docker", "k8s-prod-cluster",
             "8080,8443", "生产环境核心服务器"],
            ["web-prod-02", "10.0.1.11", "203.0.113.11", "", "keycloak-vault-02", "Linux", "Ubuntu 22.04", "5.15.0-91-generic", 8, 32, "500GB SSD",
             "北京-亦庄-A区", "支付系统", "张三", "running", "2025-01-15", "2028-01-15",
             "支付网关服务", "Spring Boot", "MySQL 8.0", "docker", "k8s-prod-cluster",
             "8080,8443", ""],
            ["db-master-01", "10.0.2.20", "", "RootPass@2024", "MySQL:root:RootPass123\nMySQL:repl_user:ReplPass456", "Linux", "CentOS 7.9", "3.10.0-1160.el7.x86_64", 32, 256, "2TB NVMe RAID10",
             "北京-亦庄-B区", "数据库服务", "李四", "running", "2024-06-01", "2027-06-01",
             "主数据库集群", "", "MySQL 8.0", "physical", "Dell R750xs",
             "3306", "主库，严禁重启"],
            ["db-slave-01", "10.0.2.21", "", "RootPass@2024", "MySQL:app_readonly:ReadOnly789", "Linux", "CentOS 7.9", "3.10.0-1160.el7.x86_64", 16, 128, "1TB SSD",
             "北京-亦庄-B区", "数据库服务", "李四", "running", "2024-06-01", "2027-06-01",
             "从数据库集群", "", "MySQL 8.0", "physical", "Dell R750xs",
             "3306", "从库-只读"],
            ["mq-01", "10.0.3.30", "", "", "", "Linux", "Ubuntu 20.04", "5.4.0-150-generic", 4, 16, "200GB SSD",
             "北京-亦庄-A区", "消息中间件", "王五", "running", "2025-03-10", "2028-03-10",
             "Kafka 集群", "", "", "vm", "VMware ESXi",
             "9092,9093", ""],
            ["redis-cluster-01", "10.0.4.40", "", "", "Redis:default:RedisPass789\nRedis:sentinel:Sentinel123", "Linux", "Ubuntu 20.04", "5.4.0-150-generic", 8, 64, "400GB SSD",
             "北京-亦庄-A区", "缓存服务", "赵六", "running", "2025-02-20", "2028-02-20",
             "Redis 集群", "", "Redis 7.0", "docker", "k8s-prod-cluster",
             "6379,16379", ""],
            ["monitor-01", "10.0.5.50", "", "", "", "Linux", "Ubuntu 22.04", "5.15.0-88-generic", 4, 16, "200GB SSD",
             "北京-亦庄-C区", "监控系统", "孙七", "running", "2025-04-01", "2028-04-01",
             "Prometheus+Grafana", "", "", "docker", "k8s-mon-cluster",
             "9090,3000", "监控+告警"],
            ["log-01", "10.0.5.51", "", "", "", "Linux", "Ubuntu 22.04", "5.15.0-88-generic", 8, 32, "1TB HDD",
             "北京-亦庄-C区", "日志平台", "孙七", "running", "2025-04-01", "2028-04-01",
             "ELK Stack", "", "Elasticsearch", "docker", "k8s-mon-cluster",
             "9200,5601", ""],
            ["backup-svr-01", "10.0.6.60", "", "Backup@2024", "NAS:admin:NASPass123", "Windows", "Windows Server 2022", "10.0.20348", 4, 16, "4TB HDD",
             "北京-亦庄-D区", "备份系统", "周八", "running", "2024-09-01", "2027-09-01",
             "定时备份服务", ".NET", "", "physical", "Dell R450",
             "445", "每周全量备份"],
            ["test-web-01", "10.0.7.70", "", "", "", "Linux", "Ubuntu 22.04", "5.15.0-78-generic", 2, 8, "100GB SSD",
             "北京-亦庄-E区", "测试环境", "吴九", "stopped", "2025-05-01", "2028-05-01",
             "测试服务", "Spring Boot", "", "vm", "VMware ESXi",
             "8080", "测试环境-已停用"],
            ["dev-api-01", "10.0.8.80", "", "", "PostgreSQL:dev_user:DevPass456", "Linux", "Debian 12", "6.1.0-13-amd64", 4, 16, "200GB SSD",
             "北京-亦庄-E区", "开发环境", "郑十", "maintenance", "2025-05-15", "2028-05-15",
             "API开发服务器", "FastAPI", "PostgreSQL", "docker", "k8s-dev-cluster",
             "8000,5432", "开发环境-维护中"],
            ["old-web-01", "192.168.1.100", "", "", "", "Windows", "Windows Server 2016", "10.0.14393", 2, 8, "100GB HDD",
             "上海-张江", "订单系统", "刘十一", "decommissioned", "2020-01-01", "2023-01-01",
             "旧版订单服务", "ASP.NET", "SQL Server", "physical", "HP DL380 G9",
             "80,443", "已报废-2025年"],
        ]
        data_rows = []
        for row in sample_rows:
            data_rows.append(dict(zip(TEMPLATE_COLUMNS, row)))
        df = pd.DataFrame(data_rows, columns=TEMPLATE_COLUMNS)
        df.columns = cn_columns
        sheet_name = "测试数据"
    else:
        df = pd.DataFrame(columns=TEMPLATE_COLUMNS)
        df.columns = cn_columns
        sheet_name = "servers"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = 16
    output.seek(0)
    return output


def validate_ip(ip: str) -> bool:
    if not ip:
        return False
    pattern = r"^(\d{1,3}\.){3}\d{1,3}$"
    return bool(re.match(pattern, ip))


def validate_date(date_str: str) -> bool:
    if not date_str or str(date_str).strip() == "":
        return True
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y", "%d/%m/%Y"]:
        try:
            datetime.strptime(str(date_str).strip(), fmt)
            return True
        except ValueError:
            continue
    return False


def import_excel(file) -> tuple[int, int, list]:
    try:
        if file.name.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file, engine="openpyxl")
    except Exception as e:
        return 0, 0, [f"文件读取错误: {str(e)}"]

    df = df.where(pd.notnull(df), "")

    # Normalize column names: strip spaces, try Chinese→English mapping
    raw_cols = [str(c).strip() for c in df.columns]
    normalized = []
    for c in raw_cols:
        if c in COLUMN_LABELS_EN:
            normalized.append(COLUMN_LABELS_EN[c])
        else:
            normalized.append(c.lower().replace(" ", "_"))
    df.columns = normalized

    valid_columns = [c for c in TEMPLATE_COLUMNS if c in df.columns]
    if not valid_columns:
        return 0, 0, ["文件中没有找到有效列"]

    success = 0
    updated = 0
    errors = []

    for idx, row in df.iterrows():
        row_num = idx + 2
        hostname = str(row.get("hostname", "")).strip()
        private_ip = str(row.get("private_ip", "")).strip()

        if not hostname:
            errors.append(f"第 {row_num} 行: hostname 为空")
            continue

        if not validate_ip(private_ip):
            errors.append(f"第 {row_num} 行: IP 格式错误 ({private_ip})")
            continue

        purchase_date = str(row.get("purchase_date", "")).strip()
        warranty_expire = str(row.get("warranty_expire", "")).strip()
        if purchase_date and not validate_date(purchase_date):
            errors.append(f"第 {row_num} 行: purchase_date 日期格式错误")
            continue
        if warranty_expire and not validate_date(warranty_expire):
            errors.append(f"第 {row_num} 行: warranty_expire 日期格式错误")
            continue

        data = {}
        for col in TEMPLATE_COLUMNS:
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            data[col] = str(val).strip()

        data["hostname"] = hostname
        data["private_ip"] = private_ip

        if "cpu_cores" in data:
            try: data["cpu_cores"] = int(float(data["cpu_cores"])) if data["cpu_cores"] else 0
            except: data["cpu_cores"] = 0
        if "memory_gb" in data:
            try: data["memory_gb"] = int(float(data["memory_gb"])) if data["memory_gb"] else 0
            except: data["memory_gb"] = 0

        try:
            conn = get_connection()
            existing = conn.execute(
                "SELECT id FROM servers WHERE private_ip = ?", (private_ip,)
            ).fetchone()
            if existing:
                update_server(existing["id"], data)
                updated += 1
            else:
                add_server(data)
                success += 1
        except Exception as e:
            errors.append(f"第 {row_num} 行: {str(e)}")

    return success, updated, errors


def export_excel(server_list: list) -> BytesIO:
    export_cols = [c for c in TEMPLATE_COLUMNS if c != "credentials"]
    df = pd.DataFrame(server_list, columns=TEMPLATE_COLUMNS)
    cols_to_export = [c for c in export_cols if c in df.columns]
    df = df[cols_to_export]
    # Use Chinese column headers for export
    rename_map = {c: COLUMN_LABELS_CN[c] for c in cols_to_export if c in COLUMN_LABELS_CN}
    df = df.rename(columns=rename_map)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="servers", index=False)
    output.seek(0)
    return output


# ===================================================================
# UI Module - CSS (Dark Theme)
# ===================================================================

def inject_css():
    st.markdown("""
    <style>
    :root {
        --bg: #0f172a;
        --card: #1e293b;
        --border: #334155;
        --text: #e2e8f0;
        --muted: #94a3b8;
        --primary: #3b82f6;
        --header-bg: #0b1120;
    }

    .stApp {
        background-color: #0f172a;
    }

    /* ===== HEADER ===== */
    .sams-header {
        background: linear-gradient(135deg, #0b1120 0%, #111f3a 100%);
        color: #e2e8f0;
        padding: 0 24px;
        height: 60px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        border-bottom: 2px solid #3b82f6;
        margin: -16px -16px 0 -16px;
    }
    .sams-header .left {
        display: flex; align-items: center; gap: 12px;
        font-size: 18px; font-weight: 700; color: #f1f5f9;
    }
    .sams-header .right {
        display: flex; align-items: center; gap: 20px;
        font-size: 13px; color: #cbd5e1;
    }
    .sams-header .logo { font-size: 24px; }

    /* ===== SIDEBAR ===== */
    section[data-testid="stSidebar"] {
        background-color: #0b1120;
    }
    section[data-testid="stSidebar"] * {
        color: #cbd5e1 !important;
    }
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3 {
        color: #f1f5f9 !important;
    }
    section[data-testid="stSidebar"] button {
        color: #cbd5e1 !important;
        background-color: transparent !important;
        border: 1px solid #1e293b !important;
        border-radius: 6px !important;
    }
    section[data-testid="stSidebar"] button:hover {
        background-color: #1e293b !important;
        border-color: #334155 !important;
        color: #e2e8f0 !important;
    }
    section[data-testid="stSidebar"] button:active,
    section[data-testid="stSidebar"] button:focus {
        background-color: #1e293b !important;
        border-color: #334155 !important;
        color: #e2e8f0 !important;
    }
    section[data-testid="stSidebar"] [data-testid="stExpander"] {
        background: transparent !important;
        border: none !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: #1e293b !important;
    }

    /* ===== STATUS BADGES ===== */
    .badge {
        display: inline-block;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 12px;
        font-weight: 600;
        color: #fff;
    }
    .badge-running { background-color: #10b981; }
    .badge-stopped { background-color: #f59e0b; color: #1f2937; }
    .badge-maintenance { background-color: #3b82f6; }
    .badge-decommissioned { background-color: #ef4444; }

    /* ===== DETAIL DRAWER ===== */
    .detail-section {
        background: #1e293b;
        border: 1px solid #334155;
        border-radius: 10px;
        padding: 16px;
    }
    .detail-row {
        display: flex;
        justify-content: space-between;
        padding: 7px 0;
        border-bottom: 1px solid #1e293b;
        font-size: 13px;
    }
    .detail-label { color: #94a3b8; font-weight: 500; }
    .detail-value { color: #e2e8f0; font-weight: 600; text-align: right; max-width: 250px; }

    /* ===== FOOTER ===== */
    .sams-footer {
        text-align: center; color: #475569; font-size: 12px;
        padding: 12px; border-top: 1px solid #1e293b; margin-top: 30px;
    }

    /* ===== INPUT FIELDS - ALWAYS VISIBLE BORDERS ===== */
    input, textarea, [data-baseweb="input"], [data-baseweb="textarea"] {
        border: 1px solid #475569 !important;
        background-color: #1e293b !important;
        color: #e2e8f0 !important;
        border-radius: 6px !important;
    }
    input:focus, textarea:focus, [data-baseweb="input"]:focus, [data-baseweb="textarea"]:focus {
        border-color: #3b82f6 !important;
        box-shadow: 0 0 0 1px #3b82f6 !important;
    }
    [data-baseweb="select"] {
        border: 1px solid #475569 !important;
        background-color: #1e293b !important;
        border-radius: 6px !important;
    }
    [data-baseweb="select"] > div {
        background-color: #1e293b !important;
        color: #e2e8f0 !important;
    }
    [data-baseweb="popover"] {
        background-color: #1e293b !important;
        border: 1px solid #475569 !important;
    }
    [data-baseweb="popover"] * {
        color: #e2e8f0 !important;
    }

    /* ===== BUTTONS ===== */
    .stButton > button {
        border-radius: 6px; font-weight: 500;
        border: 1px solid #475569 !important;
    }
    .stButton > button:hover {
        border-color: #3b82f6 !important;
    }

    /* ===== EXPANDER / FORM ===== */
    .stExpander, [data-testid="stExpander"] {
        background: #1e293b !important;
        border: 1px solid #334155 !important;
        border-radius: 10px !important;
    }
    section[data-testid="stSidebar"] [data-testid="stExpander"] {
        border: none !important;
        background: transparent !important;
    }
    [data-testid="stForm"] {
        background: transparent;
    }

    /* ===== DATAFRAME ===== */
    [data-testid="stDataFrame"] {
        border-radius: 10px; overflow: hidden;
        border: 1px solid #334155;
    }
    [data-testid="stTable"] {
        color: #e2e8f0;
    }

    /* ===== METRIC CARDS ===== */
    [data-testid="stMetric"] {
        background: #1e293b;
        border: 1px solid #334155;
        border-radius: 10px;
        padding: 16px;
    }
    [data-testid="stMetric"] label {
        color: #94a3b8 !important;
    }
    [data-testid="stMetricValue"] {
        color: #f1f5f9 !important;
        font-size: 28px !important;
    }

    /* ===== TABS ===== */
    .stTabs [data-baseweb="tab-list"] {
        border-bottom: 1px solid #334155;
    }
    .stTabs [data-baseweb="tab"] {
        color: #94a3b8;
    }
    .stTabs [aria-selected="true"] {
        color: #3b82f6;
    }

    /* ===== ALERTS ===== */
    .stAlert {
        border-radius: 8px;
    }

    /* ===== MULTISELECT ===== */
    [data-baseweb="tag"] {
        background-color: #1e3a5f !important;
        border: 1px solid #3b82f6 !important;
        color: #e2e8f0 !important;
    }

    /* ===== NUMBER INPUT ===== */
    button[data-testid="stNumberInputButton"] {
        border: 1px solid #475569 !important;
        background: #1e293b !important;
        color: #e2e8f0 !important;
    }
    </style>
    """, unsafe_allow_html=True)


# ===================================================================
# UI Module - Header
# ===================================================================

def render_header():
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db_status = "🟢" if os.path.exists(DB_PATH) else "🔴"

    st.markdown(f"""
    <div class="sams-header">
        <div class="left">
            <span class="logo">🖥</span>
            <span>服务器资产管理系统</span>
        </div>
        <div class="right">
            <span>👤 {st.session_state.get('username', 'N/A')}</span>
            <span>🕐 {current_time}</span>
            <span>DB {db_status}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ===================================================================
# UI Module - Sidebar
# ===================================================================

def render_sidebar():
    with st.sidebar:
        st.markdown("### SAMS")
        st.markdown("---")

        menu_sections = {
            "仪表盘": {
                "📊 资产总览": "dashboard",
            },
            "资产管理": {
                "📋 资产列表": "server_list",
                "➕ 新增资产": "add_server",
                "📥 批量导入": "import",
                "📤 导出资产": "export",
            },
            "统计分析": {
                "📊 资产统计分析": "analytics",
            },
            "系统管理": {
                "📝 模板下载": "template",
                "💾 数据库备份": "backup",
                "👥 用户管理": "user_mgmt",
            },
            "用户中心": {
                "👤 当前信息": "profile",
                "🔑 修改密码": "change_pwd",
                "🚪 退出登录": "logout",
            },
        }

        for section, items in menu_sections.items():
            expanded = (section == "仪表盘")
            with st.expander(section, expanded=expanded):
                for label, page in items.items():
                    if st.button(label, key=f"menu_{page}_{label}", use_container_width=True):
                        if page == "logout":
                            logout()
                            st.rerun()
                        if page == "analytics":
                            tab_map = {"业务统计": 0, "系统类型统计": 1, "运行方式统计": 2}
                            st.session_state.analytics_tab = tab_map.get(label, 0)
                        st.session_state.current_page = page
                        st.rerun()


# ===================================================================
# UI Module - Dashboard
# ===================================================================

def render_dashboard():
    st.title("仪表盘")
    stats = get_dashboard_stats()

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("总资产数", stats["total"])
    with col2:
        st.metric("运行中", stats["running"])
    with col3:
        st.metric("已停止", stats["stopped"])
    with col4:
        st.metric("维护中", stats["maintenance"])
    with col5:
        st.metric("已下线", stats["decommissioned"])

    st.markdown("---")
    st.subheader("资产状态分布")
    status_df = pd.DataFrame({
        "运行中": [stats["running"]],
        "已停止": [stats["stopped"]],
        "维护中": [stats["maintenance"]],
        "已下线": [stats["decommissioned"]],
    })
    st.bar_chart(status_df, color=["#10b981", "#f59e0b", "#3b82f6", "#ef4444"])

    st.markdown("---")
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("系统类型分布")
        if stats["os_distribution"]:
            os_df = pd.DataFrame(
                list(stats["os_distribution"].items()),
                columns=["系统类型", "数量"]
            ).set_index("系统类型")
            st.bar_chart(os_df, horizontal=True)
        else:
            st.info("暂无数据")

    with col2:
        st.subheader("运行方式分布")
        if stats["runtime_distribution"]:
            rt_df = pd.DataFrame(
                list(stats["runtime_distribution"].items()),
                columns=["运行方式", "数量"]
            ).set_index("运行方式")
            st.bar_chart(rt_df, horizontal=True)
        else:
            st.info("暂无数据")

    st.markdown("---")
    st.subheader("业务线分布")
    if stats["business_distribution"]:
        biz_df = pd.DataFrame(
            list(stats["business_distribution"].items()),
            columns=["业务线", "数量"]
        ).set_index("业务线")
        st.bar_chart(biz_df, horizontal=True)
    else:
        st.info("暂无数据")


# ===================================================================
# UI Module - Server List
# ===================================================================

def render_status_badge(status: str) -> str:
    label = STATUS_LABELS.get(status, status)
    return f'<span class="badge badge-{status}">{label}</span>'


def render_server_list():
    st.title("资产列表")

    all_businesses = get_distinct_values("business")
    all_os = get_distinct_values("os_type")
    all_runtime = get_distinct_values("runtime_type")

    search = st.text_input("🔍 搜索", placeholder="主机名 / IP地址 / 负责人 / 业务线 ...",
                           key="list_search", label_visibility="collapsed")

    with st.expander("筛选条件", expanded=False):
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            status_f = st.multiselect("状态", STATUS_OPTIONS,
                                      format_func=lambda x: STATUS_LABELS.get(x, x),
                                      key="filter_status")
        with col2:
            os_f = st.multiselect("系统类型", all_os, key="filter_os")
        with col3:
            runtime_f = st.multiselect("运行方式", all_runtime, key="filter_runtime")
        with col4:
            biz_f = st.multiselect("业务线", all_businesses, key="filter_biz")

    page = st.session_state.get("list_page", 1)
    servers, total = get_servers(
        page=page, search=search,
        status_filter=status_f if status_f else None,
        os_filter=os_f if os_f else None,
        runtime_filter=runtime_f if runtime_f else None,
        business_filter=biz_f if biz_f else None,
    )
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)

    col_left, col_right = st.columns([3, 1.5])

    with col_left:
        st.caption(f"共 {total} 条记录 | 第 {page}/{total_pages} 页")

        display_data = []
        for s in servers:
            display_data.append({
                "主机名": s["hostname"],
                "IP地址": s["private_ip"],
                "操作系统": s["os_type"],
                "CPU": s["cpu_cores"],
                "内存(GB)": s["memory_gb"],
                "业务线": s["business"],
                "负责人": s["owner"],
                "状态": STATUS_LABELS.get(s["status"], s["status"]),
                "更新时间": s["updated_at"][:10] if s["updated_at"] else "",
            })

        if display_data:
            df = pd.DataFrame(display_data)

            event = st.dataframe(
                df,
                use_container_width=True,
                height=600,
                hide_index=True,
                on_select="rerun",
                selection_mode="multi-row",
                key="server_table",
            )

            selected_ids = []
            if event is not None and hasattr(event, 'selection') and event.selection.rows:
                for idx in event.selection.rows:
                    if idx < len(servers):
                        selected_ids.append(servers[idx]["id"])
                # Single select for detail view: use last clicked
                if len(event.selection.rows) == 1:
                    st.session_state.selected_server_id = selected_ids[0]
                else:
                    st.session_state.selected_server_id = None

            # Batch delete button
            if selected_ids:
                st.markdown(f"已选中 **{len(selected_ids)}** 条资产")
                if st.button(f"🗑 批量删除 ({len(selected_ids)}条)", key="batch_del_btn", type="secondary"):
                    st.session_state.batch_delete_ids = selected_ids
                    st.rerun()

            # Batch delete confirmation
            batch_ids = st.session_state.get("batch_delete_ids", [])
            if batch_ids:
                st.error(f"⚠ 确认删除选中的 {len(batch_ids)} 条资产？此操作不可恢复。")
                cb1, cb2 = st.columns(2)
                with cb1:
                    if st.button("✅ 确认批量删除", key="confirm_batch_del", use_container_width=True):
                        for sid in batch_ids:
                            delete_server(sid)
                        st.session_state.batch_delete_ids = []
                        get_dashboard_stats.clear()
                        st.success(f"已删除 {len(batch_ids)} 条资产")
                        st.rerun()
                with cb2:
                    if st.button("❌ 取消", key="cancel_batch_del", use_container_width=True):
                        st.session_state.batch_delete_ids = []
                        st.rerun()
        else:
            st.info("没有找到匹配的资产记录")

        # Pagination
        col_p1, col_p2, col_p3, col_p4, col_p5 = st.columns([1, 1, 2, 1, 1])
        with col_p1:
            if st.button("◀ 首页", disabled=(page == 1), key="first_page"):
                st.session_state.list_page = 1
                st.rerun()
        with col_p2:
            if st.button("◂ 上一页", disabled=(page == 1), key="prev_page"):
                st.session_state.list_page = max(1, page - 1)
                st.rerun()
        with col_p3:
            go_page = st.number_input(
                "页码", min_value=1, max_value=total_pages, value=page,
                label_visibility="collapsed", key="page_input"
            )
            if go_page != page:
                st.session_state.list_page = go_page
                st.rerun()
        with col_p4:
            if st.button("下一页 ▸", disabled=(page >= total_pages), key="next_page"):
                st.session_state.list_page = min(total_pages, page + 1)
                st.rerun()
        with col_p5:
            if st.button("末页 ▶", disabled=(page >= total_pages), key="last_page"):
                st.session_state.list_page = total_pages
                st.rerun()

    # Clear selection when page changes
    last_page = st.session_state.get("_last_list_page", 0)
    if page != last_page:
        st.session_state.selected_server_id = None
        st.session_state._last_list_page = page

    with col_right:
        server_id = st.session_state.get("selected_server_id")
        if server_id:
            render_detail_drawer(server_id)


def render_detail_drawer(server_id: int):
    server = get_server_by_id(server_id)
    if not server:
        st.warning("资产不存在")
        return

    st.markdown("### 资产详情")

    fields = [
        ("主机名", "hostname"),
        ("内网IP", "private_ip"),
        ("公网IP", "public_ip"),
        ("系统密码", "credentials"),
        ("应用密码", "credential_ref"),
        ("操作系统", "os_type"),
        ("系统版本", "os_version"),
        ("内核版本", "kernel_version"),
        ("CPU(核)", "cpu_cores"),
        ("内存(GB)", "memory_gb"),
        ("磁盘", "disk_info"),
        ("位置", "location"),
        ("业务线", "business"),
        ("负责人", "owner"),
        ("状态", "status"),
        ("运行方式", "runtime_type"),
        ("运行详情", "runtime_detail"),
        ("数据库", "db_info"),
        ("端口信息", "port_info"),
        ("业务服务", "business_service"),
        ("应用框架", "app_framework"),
        ("采购日期", "purchase_date"),
        ("保修截止", "warranty_expire"),
        ("备注", "remarks"),
        ("创建时间", "created_at"),
        ("更新时间", "updated_at"),
    ]

    html_parts = ['<div class="detail-section">']
    for label, key in fields:
        val = server.get(key, "")
        if str(val) == "" or val is None:
            val = "-"
        elif key in ("credentials", "credential_ref"):
            val = "********"
        if key == "status":
            val = render_status_badge(str(val))
        html_parts.append(
            f'<div class="detail-row">'
            f'<span class="detail-label">{label}</span>'
            f'<span class="detail-value">{val}</span>'
            f'</div>'
        )
    html_parts.append('</div>')
    st.markdown("".join(html_parts), unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("✏ 编辑", key="detail_edit", use_container_width=True):
            st.session_state.edit_server_id = server_id
            st.session_state.current_page = "add_server"
            st.rerun()
    with col2:
        if st.button("🗑 删除", key="detail_delete", use_container_width=True, type="secondary"):
            st.session_state.confirm_delete_id = server_id
            st.rerun()
    with col3:
        if st.button("✖ 关闭", key="detail_close", use_container_width=True):
            st.session_state.selected_server_id = None
            st.rerun()

    if st.session_state.get("confirm_delete_id") == server_id:
        st.error("⚠ 确认删除该资产？此操作不可恢复。")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("✅ 确认删除", key="confirm_del_btn", use_container_width=True):
                delete_server(server_id)
                st.session_state.confirm_delete_id = None
                st.session_state.selected_server_id = None
                get_dashboard_stats.clear()
                st.success("已删除")
                st.rerun()
        with c2:
            if st.button("❌ 取消", key="cancel_del_btn", use_container_width=True):
                st.session_state.confirm_delete_id = None
                st.rerun()


# ===================================================================
# UI Module - Add / Edit Server
# ===================================================================

def server_form_data(defaults: dict | None = None) -> dict:
    if defaults is None:
        defaults = {}
    col1, col2 = st.columns(2)
    data = {}
    with col1:
        data["hostname"] = st.text_input("主机名 *", value=defaults.get("hostname", ""), key="f_hostname")
        data["private_ip"] = st.text_input("内网IP *", value=defaults.get("private_ip", ""), key="f_private_ip")
        data["public_ip"] = st.text_input("公网IP", value=defaults.get("public_ip", ""), key="f_public_ip")
        data["credentials"] = st.text_input("系统密码", value=defaults.get("credentials", ""), key="f_credentials", type="password",
                                            help="服务器 root/Administrator 密码")
        data["credential_ref"] = st.text_area("应用密码", value=defaults.get("credential_ref", ""), key="f_credential_ref", height=100,
                                              placeholder="应用名:用户名:密码（每行一条）\n例如:\nMySQL:root:MyPass123\nMySQL:app_user:AppPass456\nRedis:default:RedisPass789",
                                              help="格式: 应用名:用户名:密码，每行一条记录")
        data["os_type"] = st.selectbox("操作系统", [""] + OS_OPTIONS,
                                       index=([""] + OS_OPTIONS).index(defaults.get("os_type", "")) if defaults.get("os_type", "") in OS_OPTIONS else 0,
                                       key="f_os_type")
        data["os_version"] = st.text_input("系统版本", value=defaults.get("os_version", ""), key="f_os_version")
        data["kernel_version"] = st.text_input("内核版本", value=defaults.get("kernel_version", ""), key="f_kernel_version")
        data["cpu_cores"] = st.number_input("CPU(核)", min_value=0, step=1,
                                            value=int(defaults.get("cpu_cores", 0) or 0), key="f_cpu")
        data["memory_gb"] = st.number_input("内存(GB)", min_value=0, step=1,
                                            value=int(defaults.get("memory_gb", 0) or 0), key="f_memory")
        data["disk_info"] = st.text_input("磁盘", value=defaults.get("disk_info", ""), key="f_disk")
        data["location"] = st.text_input("位置", value=defaults.get("location", ""), key="f_location")
        data["business"] = st.text_input("业务线", value=defaults.get("business", ""), key="f_business")
        data["owner"] = st.text_input("负责人", value=defaults.get("owner", ""), key="f_owner")
    with col2:
        status_idx = STATUS_OPTIONS.index(defaults.get("status", "running")) if defaults.get("status", "") in STATUS_OPTIONS else 0
        data["status"] = st.selectbox("状态", STATUS_OPTIONS, index=status_idx,
                                      format_func=lambda x: STATUS_LABELS.get(x, x), key="f_status")
        rt_idx = ([""] + RUNTIME_OPTIONS).index(defaults.get("runtime_type", "")) if defaults.get("runtime_type", "") in RUNTIME_OPTIONS else 0
        data["runtime_type"] = st.selectbox("运行方式", [""] + RUNTIME_OPTIONS, index=rt_idx, key="f_runtime_type")
        data["runtime_detail"] = st.text_input("运行详情", value=defaults.get("runtime_detail", ""), key="f_runtime_detail")
        data["db_info"] = st.text_input("数据库", value=defaults.get("db_info", ""), key="f_db_info")
        data["port_info"] = st.text_input("端口信息", value=defaults.get("port_info", ""), key="f_port_info")
        data["purchase_date"] = st.text_input("采购日期", value=defaults.get("purchase_date", ""), key="f_purchase_date", placeholder="YYYY-MM-DD")
        data["warranty_expire"] = st.text_input("保修截止", value=defaults.get("warranty_expire", ""), key="f_warranty", placeholder="YYYY-MM-DD")
        data["business_service"] = st.text_input("业务服务", value=defaults.get("business_service", ""), key="f_biz_service")
        data["app_framework"] = st.text_input("应用框架", value=defaults.get("app_framework", ""), key="f_app_framework")
        data["remarks"] = st.text_area("备注", value=defaults.get("remarks", ""), key="f_remarks", height=80)
    return data


def render_add_server():
    edit_id = st.session_state.get("edit_server_id")
    existing = get_server_by_id(edit_id) if edit_id else None

    if edit_id:
        st.title("编辑资产")
        st.info(f"当前编辑: **{existing.get('hostname', 'N/A')}** ({existing.get('private_ip', 'N/A')})")
    else:
        st.title("新增资产")

    with st.form("server_form"):
        data = server_form_data(existing)

        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            submitted = st.form_submit_button("💾 保存", use_container_width=True)
        with col2:
            reset = st.form_submit_button("🔄 重置", use_container_width=True)
        with col3:
            back = st.form_submit_button("↩ 返回", use_container_width=True)

        if back:
            st.session_state.edit_server_id = None
            st.session_state.current_page = "server_list"
            st.rerun()

        if submitted:
            errors = []
            if not data["hostname"].strip():
                errors.append("主机名不能为空")
            if not validate_ip(data["private_ip"].strip()):
                errors.append("内网IP格式不正确")
            if data["purchase_date"] and not validate_date(data["purchase_date"].strip()):
                errors.append("采购日期格式不正确 (YYYY-MM-DD)")
            if data["warranty_expire"] and not validate_date(data["warranty_expire"].strip()):
                errors.append("保修截止格式不正确 (YYYY-MM-DD)")

            if errors:
                for e in errors:
                    st.error(e)
            else:
                try:
                    if edit_id:
                        update_server(edit_id, data)
                        st.success("资产更新成功")
                        st.session_state.edit_server_id = None
                    else:
                        add_server(data)
                        st.success("资产添加成功")
                    get_dashboard_stats.clear()
                    st.session_state.current_page = "server_list"
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("内网IP已存在，请检查后重试")
                except Exception as e:
                    st.error(f"保存失败: {e}")


# ===================================================================
# UI Module - Import
# ===================================================================

def render_import():
    st.title("批量导入")
    st.markdown("支持 `.xlsx` 和 `.csv` 格式，自动根据内网IP进行新增或更新。")

    uploaded_file = st.file_uploader(
        "选择文件", type=["xlsx", "csv"],
        help="模板可通过 系统管理 → 模板下载 获取"
    )

    if uploaded_file:
        with st.spinner("正在导入..."):
            success, updated, errors = import_excel(uploaded_file)

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("新增", success)
        with col2:
            st.metric("更新", updated)
        with col3:
            st.metric("失败", len(errors))

        if errors:
            st.error("失败详情")
            for e in errors[:50]:
                st.write(f"- {e}")
            if len(errors) > 50:
                st.write(f"... 还有 {len(errors) - 50} 条错误")

        if success > 0 or updated > 0:
            get_dashboard_stats.clear()
            st.success("导入完成")


# ===================================================================
# UI Module - Export
# ===================================================================

def render_export():
    st.title("导出资产")

    all_businesses = get_distinct_values("business")

    export_biz = st.multiselect("按业务线筛选（留空导出全部）", all_businesses, key="exp_biz")
    export_status = st.multiselect("按状态筛选（留空导出全部）", STATUS_OPTIONS,
                                   format_func=lambda x: STATUS_LABELS.get(x, x),
                                   key="exp_status")

    if st.button("📥 导出 Excel", use_container_width=True):
        conn = get_connection()
        conditions = []
        params = []
        if export_biz:
            conditions.append(f"business IN ({','.join(['?']*len(export_biz))})")
            params += export_biz
        if export_status:
            conditions.append(f"status IN ({','.join(['?']*len(export_status))})")
            params += export_status
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        rows = conn.execute(f"SELECT * FROM servers {where} ORDER BY updated_at DESC", params).fetchall()
        servers = [dict(r) for r in rows]

        if servers:
            excel_data = export_excel(servers)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                "💾 下载文件",
                data=excel_data,
                file_name=f"servers_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.success(f"共导出 {len(servers)} 条记录")
        else:
            st.warning("没有符合条件的数据")


# ===================================================================
# UI Module - Template Download
# ===================================================================

def render_template():
    st.title("模板下载")
    st.markdown("下载标准导入模板，用于批量导入资产数据。")

    st.markdown("### 模板字段说明")
    cn_labels = [COLUMN_LABELS_CN[c] for c in TEMPLATE_COLUMNS]
    cols_str = " | ".join(cn_labels)
    st.markdown(f"`{cols_str}`")

    st.markdown("""
    - **主机名** / **内网IP**: 必填
    - **状态**: `running` `stopped` `maintenance` `decommissioned`
    - **采购日期** / **保修截止**: YYYY-MM-DD 格式
    """)

    col1, col2 = st.columns(2)
    with col1:
        template_data = generate_template(with_data=False)
        st.download_button(
            "📥 下载空白模板",
            data=template_data,
            file_name="sams_import_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col2:
        test_data = generate_template(with_data=True)
        st.download_button(
            "📋 下载测试数据（含12条示例）",
            data=test_data,
            file_name="sams_test_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ===================================================================
# UI Module - Analytics
# ===================================================================

def render_analytics():
    st.title("统计分析")
    stats = get_dashboard_stats()

    tab_idx = st.session_state.get("analytics_tab", 0)
    tab_labels = ["业务统计", "系统类型统计", "运行方式统计"]
    selected = st.radio("选择统计维度", tab_labels, index=tab_idx, horizontal=True,
                        key=f"analytics_radio_{tab_idx}", label_visibility="collapsed")
    current_tab = tab_labels.index(selected)
    st.session_state.analytics_tab = current_tab

    if current_tab == 0:
        st.subheader("按业务线统计")
        if stats["business_distribution"]:
            biz_df = pd.DataFrame(
                list(stats["business_distribution"].items()),
                columns=["业务线", "数量"]
            ).set_index("业务线")
            st.bar_chart(biz_df, horizontal=True)
            st.dataframe(biz_df, use_container_width=True)
        else:
            st.info("暂无数据")

    elif current_tab == 1:
        st.subheader("按系统类型统计")
        if stats["os_distribution"]:
            os_df = pd.DataFrame(
                list(stats["os_distribution"].items()),
                columns=["系统类型", "数量"]
            ).set_index("系统类型")
            col1, col2 = st.columns([2, 1])
            with col1:
                st.bar_chart(os_df, horizontal=True)
            with col2:
                st.dataframe(os_df, use_container_width=True)
        else:
            st.info("暂无数据")

    elif current_tab == 2:
        st.subheader("按运行方式统计")
        if stats["runtime_distribution"]:
            rt_df = pd.DataFrame(
                list(stats["runtime_distribution"].items()),
                columns=["运行方式", "数量"]
            ).set_index("运行方式")
            col1, col2 = st.columns([2, 1])
            with col1:
                st.bar_chart(rt_df, horizontal=True)
            with col2:
                st.dataframe(rt_df, use_container_width=True)
        else:
            st.info("暂无数据")


# ===================================================================
# UI Module - Backup
# ===================================================================

def render_backup():
    st.title("数据库备份与恢复")
    st.markdown(f"**数据库路径**: `{DB_PATH}`  |  **数据库大小**: {get_db_size()}")

    st.subheader("📤 数据备份")
    if st.button("💾 立即备份", use_container_width=True):
        if os.path.exists(DB_PATH):
            backup_path = DB_PATH + f".backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            import shutil
            shutil.copy2(DB_PATH, backup_path)
            st.success(f"备份成功: `{backup_path}`")
            st.info(f"备份大小: {os.path.getsize(backup_path) / 1024:.1f} KB")
        else:
            st.warning("数据库文件不存在")

    st.markdown("---")
    st.subheader("📥 数据恢复")
    uploaded_db = st.file_uploader(
        "选择备份文件 (.db)", type=["db"],
        help="上传之前备份的 sams.db 文件恢复数据",
        key="restore_upload"
    )
    if uploaded_db:
        st.warning("⚠ 恢复将覆盖当前所有数据，不可撤销！")
        if st.button("确认恢复数据库", type="primary", use_container_width=True):
            import shutil
            if os.path.exists(DB_PATH):
                safety_path = DB_PATH + f".before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                shutil.copy2(DB_PATH, safety_path)
                st.info(f"当前数据已备份到: `{safety_path}`")
            with open(DB_PATH, "wb") as f:
                f.write(uploaded_db.getvalue())
            get_dashboard_stats.clear()
            st.success("数据库恢复成功")
            st.rerun()


# ===================================================================
# UI Module - User Management
# ===================================================================

def render_user_management():
    st.title("用户管理")

    if st.session_state.get("user_role") != "admin":
        st.error("仅管理员可访问此页面")
        return

    conn = get_connection()
    users = conn.execute("SELECT id, username, role, created_at FROM users ORDER BY id").fetchall()

    st.subheader("用户列表")
    user_data = [{"ID": u["id"], "用户名": u["username"],
                  "角色": u["role"], "创建时间": u["created_at"]} for u in users]
    st.dataframe(pd.DataFrame(user_data), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.subheader("添加用户")
    with st.form("add_user_form"):
        col1, col2 = st.columns(2)
        with col1:
            new_user = st.text_input("用户名", key="new_username")
        with col2:
            new_password = st.text_input("密码", type="password", key="new_password")
        new_role = st.selectbox("角色", ["user", "admin"], key="new_role")

        if st.form_submit_button("添加用户"):
            if not new_user or not new_password:
                st.error("用户名和密码不能为空")
            else:
                try:
                    pw_hash = hash_password(new_password)
                    conn.execute(
                        "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                        (new_user, pw_hash, new_role),
                    )
                    conn.commit()
                    st.success(f"用户 {new_user} 创建成功")
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("用户名已存在")


# ===================================================================
# UI Module - Profile & Change Password
# ===================================================================

def render_profile():
    st.title("当前用户信息")
    username = st.session_state.get("username", "N/A")
    role = st.session_state.get("user_role", "N/A")

    col1, col2 = st.columns(2)
    with col1:
        st.metric("用户名", username)
    with col2:
        st.metric("角色", role)

    st.info(f"数据库路径: {DB_PATH}")


def render_change_password():
    st.title("修改密码")

    with st.form("change_pwd_form"):
        old_pw = st.text_input("旧密码", type="password")
        new_pw = st.text_input("新密码", type="password")
        confirm_pw = st.text_input("确认新密码", type="password")

        if st.form_submit_button("确认修改"):
            if not old_pw or not new_pw:
                st.error("请填写所有字段")
            elif new_pw != confirm_pw:
                st.error("两次密码不一致")
            elif len(new_pw) < 6:
                st.error("密码长度至少6位")
            else:
                conn = get_connection()
                user = conn.execute(
                    "SELECT * FROM users WHERE username = ?",
                    (st.session_state.username,)
                ).fetchone()
                if user and verify_password(old_pw, user["password_hash"]):
                    new_hash = hash_password(new_pw)
                    conn.execute(
                        "UPDATE users SET password_hash = ? WHERE username = ?",
                        (new_hash, st.session_state.username),
                    )
                    conn.commit()
                    st.success("密码修改成功")
                else:
                    st.error("旧密码错误")


# ===================================================================
# Login Page
# ===================================================================

def render_login():
    st.markdown("<br><br>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div style="text-align: center; margin-bottom: 30px;">
            <h1 style="color: #93c5fd;">🖥 SAMS</h1>
            <p style="color: #94a3b8;">服务器资产管理系统</p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            username = st.text_input("用户名", placeholder="admin")
            password = st.text_input("密码", type="password", placeholder="admin123")
            submitted = st.form_submit_button("登 录", use_container_width=True)

            if submitted:
                if login(username, password):
                    st.session_state.current_page = "dashboard"
                    st.rerun()
                else:
                    st.error("用户名或密码错误")


# ===================================================================
# Main Router
# ===================================================================

def main():
    st.set_page_config(
        page_title="SAMS - 资产管理系统",
        page_icon="🖥",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    inject_css()

    init_db()

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "current_page" not in st.session_state:
        st.session_state.current_page = "dashboard"
    if "selected_server_id" not in st.session_state:
        st.session_state.selected_server_id = None
    if "edit_server_id" not in st.session_state:
        st.session_state.edit_server_id = None
    if "confirm_delete_id" not in st.session_state:
        st.session_state.confirm_delete_id = None
    if "list_page" not in st.session_state:
        st.session_state.list_page = 1
    if "analytics_tab" not in st.session_state:
        st.session_state.analytics_tab = 0
    if "batch_delete_ids" not in st.session_state:
        st.session_state.batch_delete_ids = []

    if not check_auth():
        render_login()
        return

    render_header()
    render_sidebar()

    page = st.session_state.current_page

    page_map = {
        "dashboard": render_dashboard,
        "server_list": render_server_list,
        "add_server": render_add_server,
        "import": render_import,
        "export": render_export,
        "template": render_template,
        "analytics": render_analytics,
        "backup": render_backup,
        "user_mgmt": render_user_management,
        "profile": render_profile,
        "change_pwd": render_change_password,
    }

    render_func = page_map.get(page, render_dashboard)
    render_func()

    st.markdown('<div class="sams-footer">SAMS v1.0 — 服务器资产管理系统</div>',
                unsafe_allow_html=True)


if __name__ == "__main__":
    main()
