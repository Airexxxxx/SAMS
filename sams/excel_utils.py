# ===================================================================
# SAMS - Excel import / export / template generation
# ===================================================================

import re
import pandas as pd
from datetime import datetime
from io import BytesIO

from sams.config import TEMPLATE_COLUMNS, COLUMN_LABELS_CN, COLUMN_LABELS_EN, EXPORT_EXCLUDED_FIELDS, ENCRYPTED_FIELDS
from sams.database import get_connection
from sams.crud import add_server, update_server
from sams.crypto_utils import encrypt_field


def generate_template(with_data: bool = False) -> BytesIO:
    from openpyxl.styles import Font, PatternFill
    cn_columns = [COLUMN_LABELS_CN[c] for c in TEMPLATE_COLUMNS]
    output = BytesIO()

    if with_data:
        sample_rows = [
            ["web-prod-01", "10.0.1.10", "203.0.113.10", "", "", "", "keycloak-vault-01", "Linux", "Ubuntu 22.04", "5.15.0-91-generic", 8, 32, "500GB SSD",
             "北京-亦庄-A区", "支付系统", "张三", "running", "2025-01-15", "2028-01-15",
             "支付网关服务", "Spring Boot", "MySQL 8.0", "docker", "k8s-prod-cluster",
             "8080,8443", "生产环境核心服务器"],
            ["web-prod-02", "10.0.1.11", "203.0.113.11", "", "", "", "keycloak-vault-02", "Linux", "Ubuntu 22.04", "5.15.0-91-generic", 8, 32, "500GB SSD",
             "北京-亦庄-A区", "支付系统", "张三", "running", "2025-01-15", "2028-01-15",
             "支付网关服务", "Spring Boot", "MySQL 8.0", "docker", "k8s-prod-cluster",
             "8080,8443", ""],
            ["db-master-01", "10.0.2.20", "", "root", "RootPass@2024", "", "MySQL:root/RootPass123\nMySQL:repl_user/ReplPass456", "Linux", "CentOS 7.9", "3.10.0-1160.el7.x86_64", 32, 256, "2TB NVMe RAID10",
             "北京-亦庄-B区", "数据库服务", "李四", "running", "2024-06-01", "2027-06-01",
             "主数据库集群", "", "MySQL 8.0", "physical", "Dell R750xs",
             "3306", "主库，严禁重启"],
            ["db-slave-01", "10.0.2.21", "", "root", "RootPass@2024", "", "MySQL:app_readonly/ReadOnly789", "Linux", "CentOS 7.9", "3.10.0-1160.el7.x86_64", 16, 128, "1TB SSD",
             "北京-亦庄-B区", "数据库服务", "李四", "running", "2024-06-01", "2027-06-01",
             "从数据库集群", "", "MySQL 8.0", "physical", "Dell R750xs",
             "3306", "从库-只读"],
            ["mq-01", "10.0.3.30", "", "", "", "", "", "Linux", "Ubuntu 20.04", "5.4.0-150-generic", 4, 16, "200GB SSD",
             "北京-亦庄-A区", "消息中间件", "王五", "running", "2025-03-10", "2028-03-10",
             "Kafka 集群", "", "", "vm", "VMware ESXi",
             "9092,9093", ""],
            ["redis-cluster-01", "10.0.4.40", "", "", "", "", "Redis:default/RedisPass789\nRedis:sentinel/Sentinel123", "Linux", "Ubuntu 20.04", "5.4.0-150-generic", 8, 64, "400GB SSD",
             "北京-亦庄-A区", "缓存服务", "赵六", "running", "2025-02-20", "2028-02-20",
             "Redis 集群", "", "Redis 7.0", "docker", "k8s-prod-cluster",
             "6379,16379", ""],
            ["monitor-01", "10.0.5.50", "", "", "", "", "", "Linux", "Ubuntu 22.04", "5.15.0-88-generic", 4, 16, "200GB SSD",
             "北京-亦庄-C区", "监控系统", "孙七", "running", "2025-04-01", "2028-04-01",
             "Prometheus+Grafana", "", "", "docker", "k8s-mon-cluster",
             "9090,3000", "监控+告警"],
            ["log-01", "10.0.5.51", "", "", "", "", "", "Linux", "Ubuntu 22.04", "5.15.0-88-generic", 8, 32, "1TB HDD",
             "北京-亦庄-C区", "日志平台", "孙七", "running", "2025-04-01", "2028-04-01",
             "ELK Stack", "", "Elasticsearch", "docker", "k8s-mon-cluster",
             "9200,5601", ""],
            ["backup-svr-01", "10.0.6.60", "", "admin", "Backup@2024", "", "NAS:admin/NASPass123", "Windows", "Windows Server 2022", "10.0.20348", 4, 16, "4TB HDD",
             "北京-亦庄-D区", "备份系统", "周八", "running", "2024-09-01", "2027-09-01",
             "定时备份服务", ".NET", "", "physical", "Dell R450",
             "445", "每周全量备份"],
            ["test-web-01", "10.0.7.70", "", "", "", "", "", "Linux", "Ubuntu 22.04", "5.15.0-78-generic", 2, 8, "100GB SSD",
             "北京-亦庄-E区", "测试环境", "吴九", "stopped", "2025-05-01", "2028-05-01",
             "测试服务", "Spring Boot", "", "vm", "VMware ESXi",
             "8080", "测试环境-已停用"],
            ["dev-api-01", "10.0.8.80", "", "", "", "", "PostgreSQL:dev_user/DevPass456", "Linux", "Debian 12", "6.1.0-13-amd64", 4, 16, "200GB SSD",
             "北京-亦庄-E区", "开发环境", "郑十", "maintenance", "2025-05-15", "2028-05-15",
             "API开发服务器", "FastAPI", "PostgreSQL", "docker", "k8s-dev-cluster",
             "8000,5432", "开发环境-维护中"],
            ["old-web-01", "192.168.1.100", "", "", "", "", "", "Windows", "Windows Server 2016", "10.0.14393", 2, 8, "100GB HDD",
             "上海-张江", "订单系统", "刘十一", "decommissioned", "2020-01-01", "2023-01-01",
             "旧版订单服务", "ASP.NET", "SQL Server", "physical", "HP DL380 G9",
             "80,443", "已报废-2025年"],
        ]
        data_rows = [dict(zip(TEMPLATE_COLUMNS, row)) for row in sample_rows]
        df = pd.DataFrame(data_rows, columns=TEMPLATE_COLUMNS)
        df.columns = cn_columns
        sheet_name = "测试数据"
    else:
        df = pd.DataFrame(columns=TEMPLATE_COLUMNS)
        df.columns = cn_columns
        sheet_name = "servers"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = 16
    output.seek(0)
    return output


def validate_ip(ip: str) -> bool:
    if not ip:
        return False
    pattern = r"^(\d{1,3}\.){3}\d{1,3}$"
    return bool(re.match(pattern, ip))


def validate_date(date_str: str) -> bool:
    if not date_str or str(date_str).strip() == "":
        return True
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y", "%d/%m/%Y"]:
        try:
            datetime.strptime(str(date_str).strip(), fmt)
            return True
        except ValueError:
            continue
    return False


def import_excel(file) -> tuple[int, int, list]:
    try:
        if file.name.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file, engine="openpyxl")
    except Exception as e:
        return 0, 0, [f"文件读取错误: {str(e)}"]

    df = df.where(pd.notnull(df), "")

    raw_cols = [str(c).strip() for c in df.columns]
    normalized = []
    for c in raw_cols:
        if c in COLUMN_LABELS_EN:
            normalized.append(COLUMN_LABELS_EN[c])
        else:
            normalized.append(c.lower().replace(" ", "_"))
    df.columns = normalized

    valid_columns = [c for c in TEMPLATE_COLUMNS if c in df.columns]
    if not valid_columns:
        return 0, 0, ["文件中没有找到有效列"]

    success = 0
    updated = 0
    errors = []

    for idx, row in df.iterrows():
        row_num = idx + 2
        hostname = str(row.get("hostname", "")).strip()
        private_ip = str(row.get("private_ip", "")).strip()

        if not hostname:
            errors.append(f"第 {row_num} 行: hostname 为空")
            continue

        if not validate_ip(private_ip):
            errors.append(f"第 {row_num} 行: IP 格式错误 ({private_ip})")
            continue

        purchase_date = str(row.get("purchase_date", "")).strip()
        warranty_expire = str(row.get("warranty_expire", "")).strip()
        if purchase_date and not validate_date(purchase_date):
            errors.append(f"第 {row_num} 行: purchase_date 日期格式错误")
            continue
        if warranty_expire and not validate_date(warranty_expire):
            errors.append(f"第 {row_num} 行: warranty_expire 日期格式错误")
            continue

        data = {}
        for col in TEMPLATE_COLUMNS:
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            data[col] = str(val).strip()

        data["hostname"] = hostname
        data["private_ip"] = private_ip

        if "cpu_cores" in data:
            try: data["cpu_cores"] = int(float(data["cpu_cores"])) if data["cpu_cores"] else 0
            except: data["cpu_cores"] = 0
        if "memory_gb" in data:
            try: data["memory_gb"] = int(float(data["memory_gb"])) if data["memory_gb"] else 0
            except: data["memory_gb"] = 0

        try:
            conn = get_connection()
            existing = conn.execute(
                "SELECT id FROM servers WHERE private_ip = ?", (private_ip,)
            ).fetchone()
            if existing:
                update_server(existing["id"], data)
                updated += 1
            else:
                add_server(data)
                success += 1
        except Exception as e:
            errors.append(f"第 {row_num} 行: {str(e)}")

    return success, updated, errors


def export_excel(server_list: list) -> BytesIO:
    export_cols = [c for c in TEMPLATE_COLUMNS if c not in EXPORT_EXCLUDED_FIELDS]
    df = pd.DataFrame(server_list, columns=TEMPLATE_COLUMNS)
    cols_to_export = [c for c in export_cols if c in df.columns]
    df = df[cols_to_export]
    rename_map = {c: COLUMN_LABELS_CN[c] for c in cols_to_export if c in COLUMN_LABELS_CN}
    df = df.rename(columns=rename_map)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="servers", index=False)
    output.seek(0)
    return output
